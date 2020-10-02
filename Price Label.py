import os
from os import startfile
from openpyxl import *
from tkinter import *
import xlrd
from openpyxl.styles import Font
from tkinter import messagebox
global name_file


def excel(): 
    name_file = "C:/Users/admin/Desktop/labels.xlsx"
    wb = load_workbook(name_file) 
    sheet = wb.active
    sheet.column_dimensions['A'].width = 10
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 30
    sheet.column_dimensions['F'].width = 40

    sheet.cell(row=1, column=1).value = "Siz"
    sheet.cell(row=1, column=2).value = "Typ"
    sheet.cell(row=1, column=3).value = "Com"
    sheet.cell(row=1, column=4).value = "Cost"
    sheet.cell(row=1, column=5).value = "Code"
    sheet.cell(row=1, column=6).value = "No.of pieces"   

def focus1(event): 
    Size_field.focus_set() 

def focus2(event):  
    Type_field.focus_set() 

def focus3(event):  
    Company_field.focus_set() 

def focus4(event): 
    Cost_field.focus_set() 

def focus5(event):
    Code_field.focus_set() 
    
def focus6(event): 
    Np_field.focus_set() 

def clear(): 
 
    Size_field.delete(0, END) 
    Cost_field.delete(0, END) 
    Code_field.delete(0, END) 
    Np_field.delete(0, END) 

def insert():

    if (Size_field.get() == "" or
        Type_field.get() == "" or
        Company_field.get() == "" or
        Cost_field.get() == "" or
        Code_field.get() == "" or
        Np_field.get() == ""): 
        messagebox.showerror("Error", "Fill all details!")
    elif((var2.get()==1 and var1.get()==1 and var3.get()==1) or (var2.get()==1 and var1.get()==1 and var3.get()==0) or (var2.get()==0 and var1.get()==1 and var3.get()==1) or (var2.get()==1 and var1.get()==0 and var3.get()==1) or (var2.get()==0 and var1.get()==0 and var3.get()==0)):
        messagebox.showerror("Error", "Mark checkbox correctly!")
    
    else:
        name_file = "C:/Users/admin/Desktop/labels.xlsx"
        wb = load_workbook(name_file) 
        sheet = wb.active
        excel()
        if(var2.get()==1):
            count = Np_field.get()
            check=Size_field.get()
            bar=Size_field.get()+Type_field.get()+Company_field.get()+Cost_field.get()
            for i in range(int(count),0,-1):
                current_row = rtrow()
                current_column = sheet.max_column
                if(check.isdigit()):
                    sheet.cell(row=current_row, column=1).value = int(check)
                else:
                    sheet.cell(row=current_row, column=1).value = check
                sheet.cell(row=current_row, column=2).value = Type_field.get() 
                sheet.cell(row=current_row, column=3).value = Company_field.get() 
                sheet.cell(row=current_row, column=4).value = int(Cost_field.get()) 
                sheet.cell(row=current_row, column=5).value = Code_field.get() 
                sheet.cell(row=current_row, column=6).value = int(Np_field.get())
                sheet.cell(row=current_row, column=7).value = bar
                sheet.cell(row=current_row, column=7).font = Font(size = 11, name = '3 of 9 Barcode')
                wb.save(name_file)  
            Size_field.focus_set()  
            clear()
            
        if(var1.get()==1):
            count = Np_field.get()
            inc = int(Inc_field.get())
            sz = Size_field.get()
            sz = sz.split('X')
            up = ((int(sz[1])-int(sz[0]))/2)+1
            siz=int(sz[0])
            cost=int(Cost_field.get())
            bar=Size_field.get()+Type_field.get()+Company_field.get()+Cost_field.get()
            for i in range(int(up),0,-1):
                for j in range(int(count),0,-1):
                    current_row = rtrow()
                    current_column = sheet.max_column 
                    sheet.cell(row=current_row, column=1).value = siz 
                    sheet.cell(row=current_row, column=2).value = Type_field.get() 
                    sheet.cell(row=current_row, column=3).value = Company_field.get() 
                    sheet.cell(row=current_row, column=4).value = cost
                    sheet.cell(row=current_row, column=5).value = Code_field.get() 
                    sheet.cell(row=current_row, column=6).value = int(Np_field.get())
                    sheet.cell(row=current_row, column=7).value = bar
                    sheet.cell(row=current_row, column=7).font = Font(size = 11, name = '3 of 9 Barcode')
                    wb.save(name_file)
                siz+=2
                cost+=inc
                  
            Size_field.focus_set()
            clear()

        if(var3.get()==1):
            count = Np_field.get()
            inc = int(Inc_field.get())
            sz = Size_field.get()
            sz = sz.split('X')
            up = ((int(sz[1])-int(sz[0])))+1
            siz=int(sz[0])
            cost=int(Cost_field.get())
            bar=Size_field.get()+Type_field.get()+Company_field.get()+Cost_field.get()
            for i in range(int(up),0,-1):
                for j in range(int(count),0,-1):
                    current_row = rtrow()
                    current_column = sheet.max_column 
                    sheet.cell(row=current_row, column=1).value = siz 
                    sheet.cell(row=current_row, column=2).value = Type_field.get() 
                    sheet.cell(row=current_row, column=3).value = Company_field.get() 
                    sheet.cell(row=current_row, column=4).value = cost
                    sheet.cell(row=current_row, column=5).value = Code_field.get() 
                    sheet.cell(row=current_row, column=6).value = int(Np_field.get())
                    sheet.cell(row=current_row, column=7).value = bar
                    sheet.cell(row=current_row, column=7).font = Font(size = 11, name = '3 of 9 Barcode')
                    wb.save(name_file)
                siz+=1
                cost+=inc
                  
            Size_field.focus_set()
            clear()
    

def rtrow():
    name_file = "C:/Users/admin/Desktop/labels.xlsx"
    wb = load_workbook(name_file) 
    sheet = wb.active
    for i in range(2,26) :                                     
        if(sheet.cell(row=i, column=1).value == None):
            return i
            break
            
def clean():
    name_file = "C:/Users/admin/Desktop/labels.xlsx"
    wb = load_workbook(name_file) 
    sheet = wb.active
    for i in range(2,26) :
        for j in range(1,7):
            sheet.cell(row=i, column=j).value = None
    wb.save(name_file)

def endjob():
    startfile("C:/Users/admin/Desktop/MV Final Stickers1.docx")

if __name__ == "__main__": 
    
    root = Tk() 
  
    root.configure(background='light green') 
 
    root.title("Price Label") 
    root.geometry("500x400") 

    heading = Label(root, text="MEGHANA VARIETIES", fg="blue", bg="light green", font='Helvetica 18 bold') 
      
    fname = Label(root, text="Data written to labels.xlsx on desktop.", bg="light green", font='Helvetica 8')    
        
    Size = Label(root, text="Size", bg="light green", font='Helvetica 12 bold') 

    Type = Label(root, text="Type", bg="light green", font='Helvetica 12 bold') 
    
    Company = Label(root, text="Company", bg="light green", font='Helvetica 12 bold') 
    
    Cost = Label(root, text="Cost", bg="light green", font='Helvetica 12 bold') 
    
    Code = Label(root, text="Code", bg="light green", font='Helvetica 12 bold')
  
    Np = Label(root, text="No.of Pieces", bg="light green", font='Helvetica 12 bold')
    
    Inc = Label(root, text="Increment", bg="light green", font='Helvetica 10 bold')
    
    Pr = Label(root, text="Processing", bg="light green", font='Helvetica 12 bold')
   
    heading.grid(row=0, column=1)
    fname.grid(row=1, column=1)
    Pr.grid(row=2, column=0)
    Size.grid(row=5, column=0) 
    Type.grid(row=6, column=0) 
    Company.grid(row=7, column=0) 
    Cost.grid(row=8, column=0)
    Inc.grid(row=9, column=0)
    Code.grid(row=10, column=0)
    Np.grid(row=11, column=0) 
  
    #fname_field = Entry(root)
    Size_field = Entry(root) 
    Type_field = Entry(root) 
    Company_field = Entry(root) 
    Cost_field = Entry(root)
    Inc_field = Entry(root)
    Code_field = Entry(root)
    Np_field = Entry(root)
    
    var1 = IntVar()
    Checkbutton(root, text="Batch", variable=var1).grid(row=3, column=0, sticky=W)
    var2 = IntVar()
    Checkbutton(root, text="Single", variable=var2).grid(row=3, column=1, sticky=W)

    var3 = IntVar()
    Checkbutton(root, text="Series", variable=var3).grid(row=3, column=2, sticky=W)
  
    Size_field.bind("<Return>", focus1) 
 
    Type_field.bind("<Return>", focus2) 

    Company_field.bind("<Return>", focus3) 

    Cost_field.bind("<Return>", focus4) 
    
    Inc_field.bind("<Return>", focus4)
    
    Code_field.bind("<Return>", focus5)
  
    Np_field.bind("<Return>", focus6) 
  
    #fname_field.grid(row=1, column=1, ipadx="100")
    Size_field.grid(row=5, column=1, ipadx="100") 
    Type_field.grid(row=6, column=1, ipadx="100") 
    Company_field.grid(row=7, column=1, ipadx="100")  
    Cost_field.grid(row=8, column=1, ipadx="100") 
    Inc_field.grid(row=9, column=1, ipadx="100")
    Code_field.grid(row=10, column=1, ipadx="100")
    Np_field.grid(row=11, column=1, ipadx="100") 
 
    endjob = Button(root, text="End Label", fg="White", 
                            bg="Magenta", command=endjob) 
    endjob.grid(row=12, column=2)
    
    submit = Button(root, text="Submit", fg="White", 
                            bg="Blue", command=insert) 
    submit.grid(row=12, column=1)
    clean = Button(root, text="Clean", fg="Black", 
                            bg="Red", command=clean) 
    clean.grid(row=2, column=1)
     
    root.mainloop()
    
